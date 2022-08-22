import openpyxl


class HomePageData:
    test_HomePage_data = [{"fullName": "Abhishek Anand", "email": "a.t.anand@gmail.com", "gender": "Male"},
                          {"fullName": "Sumit Sourav", "email": "Sourav.s@gmail.com", "gender": "Male"},
                          {"fullName": "Ram Raghunath", "email": "Ram_Raghu@gmail.com", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Abhishek\\PRACTISE\\Selenium\\PythonSelFramework\\Book1.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "test_case_name":

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
